from openpyxl import load_workbook
import re, os

def mask_content(content):
    jumin_mask = r'([0-9]{6})\s*[-–—]?\s*([1-4])([0-9]{6})'
    phone_mask = r'(\d{2,3})\s*[-–—]\s*(\d{3,4})\s*[-–—]\s*(\d{4})'
    email_mask = r'[:\s]*([a-zA-Z0-9._%+-]{2})([a-zA-Z0-9._%+-]*)@([a-zA-Z0-9.-]+\.[a-zA-Z]{2,})'
    no_find = 0
    # 주민등록번호 마스킹 (뒷자리 첫 번째 숫자는 표시)
    if re.search(jumin_mask, content):
        content = re.sub(jumin_mask, r'\1- \2******', content)
        no_find = 1
    # 전화번호 마스킹 (중간 4자리 마스킹)
    if re.search(phone_mask, content):
        content = re.sub(phone_mask, r'\1-****-\3', content)
        no_find = 1
    # 이메일 마스킹 (이름 부분에서 앞 두 글자만 남기고 마스킹)
    if re.search(email_mask, content):
        content = re.sub(email_mask, r'\1****@\3', content)
        no_find = 1
    return content, no_find

def excel_masking(file_path):
    wb = load_workbook(file_path)
    ws = wb.active

    for row in ws.iter_rows():
        for cell in row:
            cell.value, no_find = mask_content(cell.value)

    if no_find == 0:
        print("중요정보 포함되어있지 않음")
        
    # 엑셀 파일로 저장
    wb.save(f'masking_{file_path}')





if __name__ == "__main__":
    filename = 'test.xlsx'
    file_path = os.path.join("", filename)
    excel_masking(file_path)
